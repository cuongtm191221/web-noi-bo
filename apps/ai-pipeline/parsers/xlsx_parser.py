import io
from .base import ParsedDocument, PageContent


def parse_xlsx(file_bytes: bytes) -> ParsedDocument:
    """Parse XLSX with openpyxl."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    pages = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_text = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            row_str = " | ".join(str(v) if v is not None else "" for v in row)
            rows_text.append(f"Row {row_idx}: {row_str}")
        pages.append(PageContent(
            text="\n".join(rows_text),
            location={"sheet_name": sheet_name, "row_number": len(rows_text)},
        ))
    return ParsedDocument(
        pages=pages,
        metadata={"format": "xlsx", "sheet_count": len(wb.sheetnames)},
    )